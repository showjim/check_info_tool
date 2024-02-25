import openpyxl
import os
from oletools.olevba import VBA_Parser

def extract_vba(xlsm_path, output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 解析xlsm文件中的VBA宏
    vba_parser = VBA_Parser(xlsm_path)

    if vba_parser.detect_vba_macros():
        print("VBA Macros found")
        # 提取所有VBA宏代码到指定的文件夹
        for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_all_macros():

            if vba_filename.endswith('.bas') or vba_filename.endswith('.cls'):
                output_path = os.path.join(output_dir, vba_filename)
                with open(output_path, 'w') as outfile:
                    outfile.write(vba_code)
                print(f"Extracted: {output_path}")
    else:
        print("No VBA macros found.")
    # 关闭vba_parser
    vba_parser.close()


def export_sheets_to_txt(xlsm_file_path, output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Load the workbook
    workbook = openpyxl.load_workbook(xlsm_file_path, keep_vba=True)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Create a text file for the sheet
        with open(f"{sheet_name}.txt", "w", encoding="utf-8") as file:
            for row in sheet.iter_rows(values_only=True):
                # Write the row to the text file, separated by tabs
                file.write("\t".join([str(cell) if cell is not None else "" for cell in row]))
                file.write("\n")  # Newline at the end of each row

if __name__ == "__main__":
    # Call the function with the path to your .xlsm file
    export_sheets_to_txt("./Book1.xlsm", './output_directory')
    # Usage
    extract_vba('./Book1.xlsm', './output_directory')