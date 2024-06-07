import zipfile, os
import openpyxl


def unzip(file_name, path):
    unzip_format = zipfile.is_zipfile(file_name)
    if unzip_format:
        with zipfile.ZipFile(file_name, "r") as zip_file:
            for file_info in zip_file.infolist():
                file_info.filename = os.path.basename(file_info.filename)
                zip_file.extract(file_info, path + '/')
    else:
        print('Need Zip Format, Please Check!!!')


def format_str(target_str):
    if True:
        target_str = target_str.replace('=', '')
        target_str = target_str.replace('*MV', '*0.001')
        target_str = target_str.replace('*MA', '*0.001')
        target_str = target_str.replace('*UV', '*0.000001')
        target_str = target_str.replace('*UA', '*0.000001')
        target_str = target_str.replace('*V', '*1')
        target_str = target_str.replace('*A', '*1')
        target_str = target_str.replace('*PS', '*0.000000000001')
        target_str = target_str.replace('*NS', '*0.000000001')
        target_str = target_str.replace('*US', '*0.000001')
        target_str = target_str.replace('*MS', '*0.001')
        target_str = target_str.replace('*GHZ', '*1000000000')
        target_str = target_str.replace('*MHZ', '*1000000')
        target_str = target_str.replace('*KHZ', '*1000')
    else:
        target_str = target_str.replace('=', '')
        target_str = target_str.replace('*MV', '*0.001')
        target_str = target_str.replace('*MA', '*0.001')
        target_str = target_str.replace('*UV', '*0.000001')
        target_str = target_str.replace('*UA', '*0.000001')
        target_str = target_str.replace('*V', '*1')
        target_str = target_str.replace('*A', '*1')
        target_str = target_str.replace('*PS', '*0.001')
        target_str = target_str.replace('*NS', '*1')
        target_str = target_str.replace('*US', '*1000')
        target_str = target_str.replace('*MS', '*1000000')
        target_str = target_str.replace('*GHZ', '*1')
        target_str = target_str.replace('*MHZ', '*0.001')
        target_str = target_str.replace('*KHZ', '*0.000001')

    return target_str

def export_sheets_to_txt(xlsm_file_path, output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Load the workbook
    workbook = openpyxl.load_workbook(xlsm_file_path, keep_vba=True)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Create a text file for the sheet
        with open(os.path.join(output_dir, f"{sheet_name}.txt"), "w", encoding="utf-8") as file:
            for row in sheet.iter_rows(values_only=True):
                # Write the row to the text file, separated by tabs
                file.write("\t".join([str(cell) if cell is not None else "" for cell in row]))
                file.write("\n")  # Newline at the end of each row